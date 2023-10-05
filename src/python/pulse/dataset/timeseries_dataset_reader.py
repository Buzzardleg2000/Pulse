# Distributed under the Apache License, Version 2.0.
# See accompanying NOTICE file for details.

import sys
import shutil
import logging
import numbers
import tempfile
import numpy as np
from pathlib import Path
from pycel import ExcelCompiler
from dataclasses import dataclass
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import get_column_letter
from typing import Dict, List, Optional, Union

from pulse.cdm.engine import SEDataRequest, SETimeSeriesValidationTarget, SEPatientTimeSeriesValidation
from pulse.cdm.patient import SEPatient, eSex
from pulse.cdm.scalars import LengthUnit, MassUnit
from pulse.cdm.scenario import SEScenarioLog
from pulse.cdm.utils.file_utils import get_data_dir
from pulse.cdm.io.engine import (
    serialize_data_request_list_to_file,
    serialize_patient_time_series_validation_to_file
)
from pulse.cdm.io.patient import serialize_patient_from_file
from pulse.dataset.utils import generate_data_request


_pulse_logger = logging.getLogger('pulse')


def extract_patient(patient_file: Path) -> SEPatient:
    """
    Attempts to extract patient from given file. Expected
    to be a log or json patient file.

    :param patient_file: Path to patient/log file containing
        patient information.

    :raises ValueError: Unknown patient file type.

    :return: Extracted patient
    """
    if patient_file.suffix.lower() == ".json":
        p = SEPatient()
        serialize_patient_from_file(patient_file, p)
        return p
    elif patient_file.suffix.lower() == ".log":
        class SEScenarioLogPatient(SEScenarioLog):
            def __init__(self, log_file: Path):
                super().__init__(
                    log_file=log_file,
                    extract_actions=False,
                    extract_events=False
                )

                self._process_log()

            def get_patient(self):
                return self._patient

        log = SEScenarioLogPatient(log_file=patient_file)
        return log.get_patient()
    else:
        raise ValueError("Unknown patient file type: {patient_file}")


def generate_validation_targets(
    xls_file: Path,
    patient_file: Path,
    patient_validation: SEPatientTimeSeriesValidation,
    output_file: Optional[Path]=None
) -> bool:
    """
    Generates timeseries validation targets.

    :param xls_file: Path to xls file.
    :param patient_file: Path to patient/log file containing
        patient information.
    :param patient_validation: Where generated targets will be stored.
    :param output_file: (Optional) If provided, generated targets
        will be serialized to this location.

    :raises ValueError: Unknown patient file type.

    :return: Whether or not validation target generation was successful.
    """
    if not xls_file.is_file():
        _pulse_logger.error(f"Could not find xls file {xls_file}")
        return False

    if not patient_file.is_file():
        _pulse_logger.error(f"Please provide a valid patient file {patient_file}")
        return False

    # Either load patient directly from patient file or extract from log
    p = patient_validation.get_patient()
    p.copy(extract_patient(patient_file=patient_file))

    # If no name, set to filename for potential table filenames
    if not p.has_name() or not p.get_name():
        p.set_name(patient_file.stem)

    # Temp file to preserve original state through patient updates.
    # We're saving a temporary file so that pycel can get openpyxl's
    # changes. It may be possible to use pycel exclusively to avoid
    # the need for a temporary file.
    tmp_xls = tempfile.NamedTemporaryFile(suffix=".xlsx")
    tmp_xls_path = Path('./tmp.xlsx')

    # xlsx sheets to skip when generating targets and requests
    ignore_sheets = ["Patient", "CardiovascularExtended"]

    try:
        # Update patient sheet so formulas can be re-evaluated with correct parameters
        update_patient_sheet(patient=p, xls_file=xls_file, new_file=tmp_xls_path)

        # Generate targets for every xls sheet, organized by table name
        workbook = load_workbook(filename=tmp_xls_path, data_only=False)
        evaluator = ExcelCompiler(filename=tmp_xls_path)
        for system in workbook.sheetnames:
            if system in ignore_sheets:
                continue
            if not generate_sheet_targets(
                sheet=workbook[system],
                evaluator=evaluator,
                patient_validation=patient_validation
            ):
                _pulse_logger.error(f"Unable to generate targets for {system} sheet")

        # Serialize generated targets, if requested
        if output_file is not None:
            _pulse_logger.info(f"Writing {output_file}")
            output_file.parent.mkdir(parents=True, exist_ok=True)
            serialize_patient_time_series_validation_to_file(patient_validation, output_file)
    except Exception as e:
        raise
    finally:  # Always clean up temp file
        tmp_xls.close()

    return True

def update_patient_sheet(patient: SEPatient, xls_file: Path, new_file: Optional[Path]=None) -> None:
    """
    Updates workbook with given patient information.

    :param patient: Relevant patient.
    :param xls_file: Original xls_file path.
    :param new_file: (Optional) If provided, the original file will remain
        unchanged and edits will be saved to this file.
    """
    workbook = load_workbook(filename=xls_file)
    sheet = workbook["Patient"]
    sheet["C2"] = "Male" if patient.get_sex() == eSex.Male else "Female"
    if patient.has_height():
        sheet["C3"] = patient.get_height().get_value(units=LengthUnit.cm)
    if patient.has_weight():
        sheet["C4"] = patient.get_weight().get_value(units=MassUnit.kg)
    if patient.has_right_lung_ratio():
        sheet["C9"] = patient.get_right_lung_ratio().get_value()
    if patient.has_body_fat_fraction():
        sheet["C12"] = patient.get_body_fat_fraction().get_value()

    if new_file is None:
        new_file = xls_file
    workbook.save(filename=new_file)


def generate_data_requests(xls_file: Path, output_dir: Path) -> None:
    """
    Generates all data requests for given xls file.

    :param xls_file: Path to xls_file containing data request information.
    :param output_dir: Path to directory to save generated data requests.
    """
    try:
        if output_dir.is_dir():
            shutil.rmtree(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        _pulse_logger.error(f"Unable to clean directories")
        return

    # Generate data requests for every xls sheet, organized by table name
    ignore_sheets = ["Patient"]
    workbook = load_workbook(filename=xls_file, data_only=True)
    dr_dict = dict()
    for system in workbook.sheetnames:
        if system in ignore_sheets:
            continue
        if not generate_sheet_requests(
                sheet=workbook[system],
                dr_dict=dr_dict
        ):
            _pulse_logger.error(f"Unable to generate requests for {system} sheet")

    # Serialize data request files (one for each table name)
    for dr_file, drs in dr_dict.items():
        filename = output_dir / f"{dr_file}.json"
        _pulse_logger.info(f"Writing {filename}")
        serialize_data_request_list_to_file(drs, filename)


def generate_sheet_requests(sheet: Worksheet, dr_dict: Dict[str, List[SEDataRequest]]) -> bool:
    """
    Generates data requests from given worksheet.

    :param sheet: Relevant xls worksheet.
    :param dr_dict: Dictionary to store generated data requests. Key is table name.

    :return: Whether generating data requests for this sheet was successful.
    """
    system = sheet.title.replace(" ", "")

    # Get header to dataclass mapping
    ws_headers = [cell.value for cell in sheet[1]]
    try:
        DRB_HEADER = ws_headers.index('Output')
        DRB_UNITS = ws_headers.index('Units')
        DRB_TABLE = ws_headers.index('Table')
        DRB_REQUEST_TYPE = ws_headers.index('Request Type')
        DRB_REQUEST_PRECISION = ws_headers.index('Request Precision')
    except ValueError as e:
        _pulse_logger.error(f"Missing required header {str(e)[:str(e).find(' is not in list')]}")
        return False

    @dataclass
    class DataRequestBuilder:
        header: str
        units: str
        dr_file: str
        request_type: str
        precision: Union[str, numbers.Number]

    for row_num, r in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        drb = DataRequestBuilder(
            header=r[DRB_HEADER],
            units=r[DRB_UNITS] if r[DRB_UNITS] else "unitless",
            dr_file=r[DRB_TABLE] if r[DRB_TABLE] else "Orphaned",
            request_type=r[DRB_REQUEST_TYPE],
            precision=r[DRB_REQUEST_PRECISION]
        )

        if not drb.header:
            continue

        # Skip header rows
        if drb.dr_file == "Table":
            continue

        if not drb.request_type:
            _pulse_logger.info(f"Not generating request {drb.header} (no DR type)")
            continue

        if "@" in drb.request_type:
            # Currently ignoring assessment types
            continue

        if drb.dr_file not in dr_dict:
            dr_dict[drb.dr_file] = []
        drs = dr_dict[drb.dr_file]

        dr = generate_data_request(
            request_type=drb.request_type,
            property_name=drb.header.replace("*", ""),
            unit_str=drb.units.strip(),
            precision=int(drb.precision) if drb.precision else None
        )

        drs.append(dr)

    return True


def generate_sheet_targets(
    sheet: Worksheet,
    evaluator: ExcelCompiler,
    patient_validation: SEPatientTimeSeriesValidation
) -> bool:
    """
    Generates validation targets from given worksheet.

    :param sheet: Relevant xls worksheet.
    :param evaluator: xls evaluator to retrieve computed cell values.
    :param patient_validation: Where generated validation targets will be stored.

    :return: Whether generating validation targets for this sheet was successful.
    """
    system = sheet.title
    targets = patient_validation.get_targets()

    # Get header to dataclass mapping
    ws_headers = [cell.value for cell in sheet[1]]
    try:
        VTB_HEADER = ws_headers.index('Output')
        VTB_UNITS = ws_headers.index('Units')
        VTB_ALGO = ws_headers.index('Algorithm')
        VTB_REF_CELL = ws_headers.index('Reference Values')
        VTB_REFS = ws_headers.index('References')
        VTB_NOTES = ws_headers.index('Notes')
        VTB_TGT_DEST = ws_headers.index('Table')
        VTB_REQUEST_TYPE = ws_headers.index('Request Type')
        VTB_TABLE_PRECISION = ws_headers.index('Table Precision')
        VTB_PATIENT_SPECIFIC = ws_headers.index('PatientSpecific')
    except ValueError as e:
        _pulse_logger.error(f"Missing required header {str(e)[:str(e).find(' is not in list')]}")
        return False

    @dataclass
    class ValidationTargetBuilder():
        header: str
        units: str
        algorithm: str
        ref_cell: Union[str, numbers.Number]
        tgt_dest: str
        request_type: str
        references: str
        notes: str
        table_precision: str
        patient_specific: bool

    for row_num, r in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        vtb = ValidationTargetBuilder(
            header=r[VTB_HEADER],
            units=r[VTB_UNITS] if r[VTB_UNITS] else "unitless",
            algorithm=r[VTB_ALGO],
            ref_cell=r[VTB_REF_CELL],
            tgt_dest=r[VTB_TGT_DEST] if r[VTB_TGT_DEST] else "Orphaned",
            request_type=r[VTB_REQUEST_TYPE],
            references=r[VTB_REFS] if r[VTB_REFS] else "",
            notes=r[VTB_NOTES] if r[VTB_NOTES] else "",
            table_precision=f".{r[VTB_TABLE_PRECISION]}" if r[VTB_TABLE_PRECISION] else "",
            patient_specific=True if r[VTB_PATIENT_SPECIFIC] and r[VTB_PATIENT_SPECIFIC].lower() == "y" else False
        )
        if not vtb.header:
            continue

        # Skip header
        if vtb.tgt_dest == "Table":
            continue

        # Nothing to validate to
        if '*' in vtb.header:
            # _pulse_logger.info(f"Ignoring request {vtb.header} (has asterisks)")
            continue
        if vtb.ref_cell is None:
            _pulse_logger.info(f"Not validating {vtb.header} (no reference value)")
            continue

        # Only generate targets with provided DR type
        if not vtb.request_type:
            _pulse_logger.info(f"Not validating {vtb.header} (no DR type)")
            continue
        if "@" in vtb.request_type:
            # Currently ignoring assessment types
            continue

        if vtb.tgt_dest not in targets:
            targets[vtb.tgt_dest] = list()
        vts = targets[vtb.tgt_dest]

        dr = generate_data_request(
            request_type=vtb.request_type,
            property_name=vtb.header,
            unit_str=vtb.units.strip(),
            precision=None
        )
        tgt = SETimeSeriesValidationTarget()
        tgt.set_header(dr.to_string())
        tgt.set_reference(vtb.references)
        tgt.set_notes(vtb.notes)
        tgt.set_patient_specific_setting(vtb.patient_specific)
        tgt.set_table_formatting(vtb.table_precision)

        ref_val = vtb.ref_cell

        algo = vtb.algorithm
        if algo == "Max":
            algo = "Maximum"
        elif algo == "Min":
            algo = "Minimum"
        elif algo.endswith("(kg)"):
            algo = algo.replace("(kg)", "_kg")
        try:
            algo = SETimeSeriesValidationTarget.eTargetType[algo]
        except:
            _pulse_logger.error(f"Unknown validation target type: {algo}. Skipping validation target for {vtb.header} in {system}.")
            continue

        # Evaluate cell if needed
        if isinstance(ref_val, str) and ref_val.startswith("="):
            if ref_val.startswith("="):
                cell_loc = f"{system}!{get_column_letter(VTB_REF_CELL+1)}{row_num+2}"
                ref_val = evaluator.evaluate(cell_loc)

        # TODO: Support other comparison types?
        if isinstance(ref_val, str):
            s_vals = ref_val.strip()
            s_vals = s_vals.replace('[', ' ')
            s_vals = s_vals.replace(']', ' ')
            vals = [float(s) for s in s_vals.split(',')]
            tgt.set_range(min(vals), max(vals), algo)
        elif isinstance(ref_val, numbers.Number):
            val = float(ref_val)
            tgt.set_equal_to(val, algo)
        else:
            _pulse_logger.warning(f"Unknown reference value type {ref_val}")
            tgt.set_range_min(np.nan)
            tgt.set_range_max(np.nan)

        vts.append(tgt)

    return True


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
    logging.getLogger("pycel").setLevel(logging.WARNING)

    xls_file = None

    if len(sys.argv) < 2:
        _pulse_logger.error("Expected inputs : <xls validation file path>")
        sys.exit(1)

    xls_file = Path(sys.argv[1])
    if not xls_file.is_file():
        xls_file = Path(get_data_dir()+sys.argv[1])
        if not xls_file.is_file():
            _pulse_logger.error("Please provide a valid xls file")
            sys.exit(1)

    output_dir = Path("./validation/")
    xls_basename = "".join(xls_file.name.rsplit("".join(xls_file.suffixes), 1))
    xls_basename_out = xls_basename[:-4] if xls_basename.lower().endswith("data") else xls_basename
    output_dir = output_dir / xls_basename_out
    try:
        if output_dir.is_dir():
            shutil.rmtree(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
    except OSError as e:
        _pulse_logger.error(f"Unable to clean directories")
        sys.exit(1)

    # Generate system validation targets for every patient file
    patient_dir = Path("./patients/")
    if patient_dir.is_dir():
        for patient_file in patient_dir.glob("*.json"):
            full_output_path = output_dir / f"{patient_file.stem}_SystemTargets.json"
            full_output_path.parent.mkdir(parents=True, exist_ok=True)

            if not generate_validation_targets(
                xls_file=xls_file,
                patient_file=patient_file,
                output_file=full_output_path
            ):
                _pulse_logger.error(f"Failed to generate validation targets for {patient_file}")
    else:
        _pulse_logger.error("Unable to locate patient directory")
