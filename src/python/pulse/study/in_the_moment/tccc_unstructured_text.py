# Distributed under the Apache License, Version 2.0.
# See accompanying NOTICE file for details.

import re
import sys
import json
import random
import logging
from enum import Enum
from pathlib import Path
from string import Template
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Any, Dict, Hashable, Iterable, List, NamedTuple, Optional, Tuple

from pulse.cdm.engine import SEEventChange
from pulse.cdm.scalars import TimeUnit, SEScalarTime
from pulse.cdm.scenario import SEObservationReportModule
from pulse.study.in_the_moment.unstructured_text import SEUnstructuredTextProperty, SEUnstructuredTextGroup
from pulse.study.in_the_moment.io.unstructured_text import serialize_unstructured_text_corpus_to_file, \
                                                           serialize_unstructured_text_group_to_file, \
                                                           serialize_unstructured_text_corpus_from_file, \
                                                           serialize_unstructured_text_group_from_file


_pulse_logger = logging.getLogger('pulse')


class TCCCUnstructuredTextReader:
    __slots__ = ("_xls_file", "_corpus", "_elapsed_time")

    def __init__(self, xls_file: Path):
        self._xls_file = xls_file

    def _process_elapsed_time(self, sheet: Worksheet) -> bool:
        """
        Parse elapsed time phrases from given spreadsheet.
        """
        phrases = list()
        for r in sheet.iter_rows(min_row=2, values_only=True):
            phrases.append(Template(r[0]))

        time_property = SEUnstructuredTextProperty()
        time_property.set_property("Time(s)")
        time_property.set_property_type(SEUnstructuredTextProperty.ePropertyType.DataRequest)
        time_property.set_any_value()

        self._elapsed_time = SEUnstructuredTextGroup()
        self._elapsed_time.set_properties([time_property])
        self._elapsed_time.set_phrases([phrases])

        return True

    def _process_vitals(self, sheet: Worksheet) -> bool:
        """
        Parse vital-based phrases from given spreadsheet.
        """
        class Stage(Enum):
            IDTableStart  = 0
            DataRequest   = 1
            CollectValues = 2

        dr_string = None
        stage = Stage.IDTableStart
        skip = False
        for r in sheet.iter_rows(min_row=1, values_only=True):
            if stage == Stage.IDTableStart:
                if all(item is None for item in r):  # Empty row
                    continue

                # First non-empty row of table is ignored
                stage = Stage.DataRequest
                continue
            elif stage == Stage.DataRequest:
                dr_string = r[0].strip()
                if dr_string.endswith("*"):
                    skip = True
                stage = Stage.CollectValues
                continue
            elif stage == Stage.CollectValues:
                if all(item is None for item in r):  # Empty row
                    stage = Stage.IDTableStart
                    skip = False
                    continue
                elif not skip:
                    compare_values = []
                    comparator_string = None
                    for match in re.finditer(r"\d+", r[0]):
                        compare_values.append(float(r[0][match.start():match.end()].strip()))
                        if comparator_string is None:
                            comparator_string = r[0][:match.start()].strip()

                    text_prop = SEUnstructuredTextProperty()
                    text_prop.set_property(dr_string)
                    text_prop.set_property_type(SEUnstructuredTextProperty.ePropertyType.DataRequest)
                    min_inclusive = False
                    max_inclusive = False
                    if len(compare_values) == 2:
                        if r[0].strip()[0] == "[":
                            min_inclusive = True
                        if r[0].strip()[-1] == "]":
                            max_inclusive = True
                        text_prop.set_range(
                            min=compare_values[0],
                            max=compare_values[1],
                            min_inclusive=min_inclusive,
                            max_inclusive=max_inclusive
                        )
                    elif len(compare_values) == 1:
                        compare_fn = {
                            ">": text_prop.set_gt,
                            ">=": text_prop.set_ge,
                            "<": text_prop.set_lt,
                            "<=": text_prop.set_le,
                            "==": text_prop.set_eq
                        }
                        if comparator_string not in compare_fn:
                            raise ValueError(f"Unknown comparator {comparator_string}")

                        compare_fn[comparator_string](compare_values[0])
                    else:
                        raise ValueError(f"Invalid comparison {r[0]}")

                    text_group = SEUnstructuredTextGroup()
                    text_group.set_properties([text_prop])
                    text_group.set_phrases([[Template(phrase.strip()) for phrase in r[1].splitlines()]])
                    self._corpus.append(text_group)

                continue

        return True

    def parse(self, corpus_file: Path, elapsed_time_file: Path) -> None:
        """
        Parse corpus from relevant xlsx file and serialize to given files.

        Elapsed time is kept separate as it is stateful.
        """
        _pulse_logger.info(f"Generating corpus from {self._xls_file}")

        self._elapsed_time = None
        self._corpus = list()

        workbook = load_workbook(filename=self._xls_file, data_only=True)

        processors = {
            "Duration From Injury": self._process_elapsed_time,
            "Vitals": self._process_vitals
        }
        for processor_sheet, processor in processors.items():
            if processor_sheet in workbook.sheetnames:
                if not processor(workbook[processor_sheet]):
                    _pulse_logger.error(f"Unable to read {processor_sheet} sheet")
            else:
                _pulse_logger.warning(f"No {processor_sheet} sheet present")

        corpus_file.parent.mkdir(parents=True, exist_ok=True)
        serialize_unstructured_text_corpus_to_file(self._corpus, corpus_file)
        if self._elapsed_time is not None:
            elapsed_time_file.parent.mkdir(parents=True, exist_ok=True)
            serialize_unstructured_text_group_to_file(self._elapsed_time, elapsed_time_file)


class TCCCUnstructuredText(SEObservationReportModule):
    __slots__ = ("_seed", "_injury_time", "_elapsed_time", "_corpus", "_events", "_reported_vitals")

    TIME_s = "Time(s)"
    RR_Per_min = "RespirationRate(1/min)"
    HR_Per_min = "HeartRate(1/min)"

    def __init__(
        self,
        corpus_json: Path,
        elapsed_time_json: Path,
        reported_vitals: List[str]=None,
        seed: Optional[int]=None
    ):
        self._headers = [
            self.TIME_s,
            self.RR_Per_min,
            self.HR_Per_min
        ]

        self._seed = seed
        if self._seed is not None:
            random.seed(self._seed)

        self._injury_time = None
        self._events = list()

        # Read in pre-parsed corpus
        self._corpus = list()
        self._elapsed_time = SEUnstructuredTextGroup()
        serialize_unstructured_text_corpus_from_file(corpus_json, self._corpus)
        serialize_unstructured_text_group_from_file(elapsed_time_json, self._elapsed_time)

        if reported_vitals is None:
            reported_vitals = list()
        self._reported_vitals = reported_vitals

        self._headers.extend(h for h in self._reported_vitals if h not in self._headers)

    def _add_phrases(self, choices: List[List[Template]], out_phrases: List[Template]) -> None:
        """
        Add a random phrase from each list of given choices to list of phrases.
        """
        for c in choices:
            out_phrases.append(random.choice(c))

    def handle_event(self, change: SEEventChange) -> None:
        """ Process given event change """
        # TODO: Handle events

    def handle_action(self, action: str, action_time: SEScalarTime) -> None:
        """ Process action """
        # Load into dict direct from JSON because we can't serialize actions from bind
        action_data = json.loads(action)

        # TCCC handles a restricted set of actions:
        # Hemorrhage, AirwayObstruction, TensionPneumothorax, NeedleDecompression, ChestOcclusiveDressing
        # Certain actions with 0 severity are being treated as interventions (such as hemorrhage -> tourniquet)
        PATIENT_ACTION = "PatientAction"
        if PATIENT_ACTION in action_data:
            action_name = list(action_data[PATIENT_ACTION].keys())[0]

            severity = 0
            insult = False
            if "Severity" in action_data[PATIENT_ACTION][action_name]:
                insult = True
                severity = action_data[PATIENT_ACTION][action_name]["Severity"]["Scalar0To1"]["Value"]

            if severity == 0:  # Severity 0 indicates intervention
                insult = False
            elif self._injury_time is None:  # First insult time
                self._injury_time = action_time


    def update(self, data_slice: NamedTuple, slice_idx: Dict[str, int]) -> Iterable[Tuple[Hashable, Any]]:
        """
        Generate unstructured text from current vitals and handled actions and events.
        """
        phrases = list()
        template_vals = dict()

        # Elapsed time is a stateful property
        if self._injury_time is not None:
            time_elapsed_s = data_slice[slice_idx[self.TIME_s]] - self._injury_time.get_value(TimeUnit.s)
            time_elapsed_min = int(round(SEScalarTime(time_elapsed_s, TimeUnit.s).get_value(TimeUnit.min)))

            template_vals["elapsed_time_m"] = f"{time_elapsed_min} " \
                                              f"{'minutes' if time_elapsed_min != 1 else 'minute'}"
            template_vals["elapsed_time_verb"] = "have" if time_elapsed_min != 1 else "has"

            self._add_phrases(choices=self._elapsed_time.get_phrases(), out_phrases=phrases)

        for vital in self._reported_vitals:
            phrases.append(Template(f"{vital} is {round(data_slice[slice_idx[vital]])}."))

        # Process stateless properties
        for property_group in self._corpus:
            if property_group.evaluate(
                data_slice=data_slice,
                slice_idx=slice_idx,
                events=self._events
            ):
                self._add_phrases(choices=property_group.get_phrases(), out_phrases=phrases)

        # Remove empty phrases, substitute template values, and shuffle
        phrases = [phrase.safe_substitute(template_vals) for phrase in phrases if phrase]
        random.shuffle(phrases)

        return [
            ("UnstructuredText", " ".join(phrases))
        ]

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

    if len(sys.argv) < 1:
        _pulse_logger.error("Expected inputs : <xls file>")
        sys.exit(1)

    xls_file = Path(sys.argv[1])
    if not xls_file.is_file():
        _pulse_logger.error("Please provide a valid xls file")
        sys.exit(1)


    corpus_json = Path(f"{xls_file.parent}/{xls_file.stem}_corpus.json")
    elapsed_time_json = Path(f"{xls_file.parent}/{xls_file.stem}_elapsed_time.json")

    reader = TCCCUnstructuredTextReader(xls_file=xls_file)
    reader.parse(corpus_file=corpus_json, elapsed_time_file=elapsed_time_json)
